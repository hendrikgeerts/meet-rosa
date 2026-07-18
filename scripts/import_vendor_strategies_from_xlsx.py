"""Bulk-importeer vendor_strategies uit you's review-spreadsheet.

Gebruik na elke kwartaal-run waarin je `nog-uit-te-zoeken.csv` met de
"hoe_gevonden" kolom hebt ingevuld. Het script herkent vier patronen:

  - "negeren" / "verkeerde betaling" / "niks mee doen"  → source_kind=ignore
  - "fysieke bon" / "fysieke factuur"                   → source_kind=physical
  - "via portaal" / "stuurt geen factuur"               → source_kind=portal
  - "vanuit foo@bar.com" / "via no-reply@..."           → source_kind=email
                                                          + email_query_hint

Vendors zonder herkenbare hint worden geskipt (zichtbaar in dry-run).

Usage:
    PYTHONPATH=src ./venv/bin/python scripts/import_vendor_strategies_from_xlsx.py \\
        --xlsx path/to/vendor-feedback.xlsx --dry-run
    PYTHONPATH=src ./venv/bin/python scripts/import_vendor_strategies_from_xlsx.py \\
        --xlsx path/to/vendor-feedback.xlsx
"""
from __future__ import annotations

import argparse
import re
import sqlite3
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from core.config import load_settings
from extensions.receipt_collector.matcher import _clean_vendor_for_search
from extensions.receipt_collector.schema import (
    find_vendor_strategy, upsert_vendor_strategy,
)
from openpyxl import load_workbook


_EMAIL_RE = re.compile(r"\b([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,})\b")

# Patroon-volgorde is significant: ignore wint van portal wint van physical
# wint van email-detectie. Eerste match is leidend.
PATTERNS = [
    ("ignore",   ["negeren", "niks mee doen", "verkeerde betaling",
                   "geen factuur, test"]),
    ("physical", ["fysieke bon", "fysieke factuur", "fysieke kassabon"]),
    ("portal",   ["via portaal", "uit portaal", "in portaal",
                   "vanuit portaal", "uit portal", "in portal",
                   "via portal", "stuurt geen factuur",
                   "stuurt geen facturen", "moeten dus altijd via portal"]),
]


def classify(hoe: str) -> str | None:
    text = hoe.lower()
    for kind, keywords in PATTERNS:
        if any(k in text for k in keywords):
            return kind
    if _EMAIL_RE.search(hoe):
        return "email"
    return None


def extract_email_hint(hoe: str) -> str | None:
    # Own domains uit settings zodat forwards van jezelf niet als
    # vendor-email worden opgepakt.
    own_domains: set[str] = set()
    try:
        from core.config import load_settings
        own_domains = set(load_settings().own_email_domains or ())
    except Exception:
        pass

    matches = _EMAIL_RE.findall(hoe)
    for m in matches:
        domain = m.split("@", 1)[-1].lower()
        if domain in own_domains:
            continue
        return f"from:{m.lower()}"
    return None


def derive_strategy(vendor_raw: str, hoe: str) -> dict | None:
    kind = classify(hoe)
    if kind is None:
        return None
    cleaned = _clean_vendor_for_search(vendor_raw).strip()
    name = cleaned or vendor_raw[:40]
    aliases = [vendor_raw]
    if cleaned and cleaned.lower() != vendor_raw.lower():
        aliases.append(cleaned)
    strat = {
        "vendor_raw": vendor_raw,
        "name": name,
        "source_kind": kind,
        "aliases": aliases,
        "email_query_hint": None,
        "portal_notes": hoe.strip()[:300],
    }
    if kind == "email":
        hint = extract_email_hint(hoe)
        if not hint:
            return None  # email-kind without parseable address — skip
        strat["email_query_hint"] = hint
    return strat


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.is_absolute():
        xlsx = ROOT / xlsx
    if not xlsx.exists():
        print(f"file not found: {xlsx}", file=sys.stderr)
        return 1

    wb = load_workbook(xlsx, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    try:
        vendor_col = next(i for i, h in enumerate(header)
                           if h and "vendor_raw" in str(h).lower())
        hoe_col = next(i for i, h in enumerate(header)
                        if h and "hoe_gevonden" in str(h).lower())
    except StopIteration:
        print("vendor_raw + hoe_gevonden kolom niet gevonden", file=sys.stderr)
        return 1

    seen: set[str] = set()
    suggestions = []
    skipped_no_hint = 0
    skipped_empty = 0
    for row in rows[1:]:
        vendor_raw = str(row[vendor_col] or "").strip()
        hoe = str(row[hoe_col] or "").strip()
        if not vendor_raw or not hoe:
            skipped_empty += 1
            continue
        if vendor_raw in seen:
            continue
        seen.add(vendor_raw)
        strat = derive_strategy(vendor_raw, hoe)
        if strat is None:
            skipped_no_hint += 1
            continue
        suggestions.append(strat)

    by_kind: dict[str, int] = {}
    for s in suggestions:
        by_kind[s["source_kind"]] = by_kind.get(s["source_kind"], 0) + 1
    print(f"\ngevonden strategies: {len(suggestions)}  (per kind: {by_kind})")
    print(f"skipped: {skipped_no_hint} zonder herkenbaar patroon, "
           f"{skipped_empty} lege rijen")

    print("\nDetails:")
    for s in suggestions:
        hint_str = (f"  hint={s['email_query_hint']}"
                    if s["email_query_hint"] else "")
        print(f"  [{s['source_kind']:8}] {s['name'][:35]:35}{hint_str}")

    if args.dry_run:
        print(f"\n[dry-run] {len(suggestions)} would be upserted")
        return 0

    settings = load_settings()
    written = 0
    overwritten = 0
    with sqlite3.connect(settings.db_path, isolation_level=None) as c:
        for s in suggestions:
            existing = find_vendor_strategy(c, vendor_text=s["vendor_raw"])
            if existing and existing.get("email_query_hint") and not s["email_query_hint"]:
                # bestaande email-hint niet overschrijven met portal/physical/ignore
                continue
            if existing:
                # Behoud bestaande naam zodat we GEEN nieuwe rij naast de
                # oude maken (root cause Run-9 Odido/CCV/Van-der-Valk
                # duplicate-bug). Merge aliases zodat alle vendor_raw
                # varianten gedekt blijven.
                target_name = existing["name"]
                merged_aliases = sorted(set(
                    list(existing.get("aliases") or []) + s["aliases"]
                ))
                merged_hint = (s.get("email_query_hint")
                                or existing.get("email_query_hint"))
                upsert_vendor_strategy(
                    c, name=target_name,
                    source_kind=s["source_kind"],
                    aliases=merged_aliases,
                    email_query_hint=merged_hint,
                    portal_notes=s["portal_notes"],
                )
                overwritten += 1
            else:
                upsert_vendor_strategy(
                    c, name=s["name"],
                    source_kind=s["source_kind"],
                    aliases=s["aliases"],
                    email_query_hint=s["email_query_hint"],
                    portal_notes=s["portal_notes"],
                )
                written += 1
    print(f"\nupserted: +{written} new, {overwritten} updated in-place")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
