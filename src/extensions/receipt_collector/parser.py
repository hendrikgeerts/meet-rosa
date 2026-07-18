"""Excel parser voor afschrijvings-lijsten.

Heuristisch: zoekt date/amount/vendor kolommen op basis van header-namen
in NL/EN. Meeste bank/admin-exports hebben varianten als "Datum"/"Date",
"Bedrag"/"Amount"/"Mutatie", "Naam"/"Tegenrekening"/"Omschrijving".

Geeft een lijst `Transaction` terug. Onleesbare rijen worden overgeslagen
met log-warning (geen exception — een typo in 1 rij mag niet de hele
batch slopen).
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

log = logging.getLogger(__name__)


# Header-aliases per logisch veld. Match-case-insensitief op gestripte tekst.
_DATE_HEADERS = {
    "datum", "date", "boekdatum", "transactiedatum", "valutadatum",
    "afschrijfdatum", "transaction date", "trans. date",
}
_AMOUNT_HEADERS = {
    "bedrag", "amount", "mutatie", "bedrag (eur)", "bedrag eur",
    "totaal", "total", "amount (eur)", "debet", "credit",
    "bij/af", "af bij", "bedrag €",
}
# Aparte set voor exports met gesplitste in/uit kolommen (bv. boekhouders)
_AMOUNT_OUT_HEADERS = {"bedrag uit", "uit", "af", "debet", "debit", "expense"}
_AMOUNT_IN_HEADERS = {"bedrag in", "in", "bij", "credit", "income"}
_VENDOR_HEADERS = {
    "naam", "vendor", "tegenrekening", "tegenpartij", "leverancier",
    "begunstigde", "naam tegenrekening", "name", "merchant", "beneficiary",
    "from", "name / description", "crediteur",
}
_DESC_HEADERS = {
    "omschrijving", "description", "mededelingen", "memo", "notes",
    "bestand", "details", "boodschap",
}
_DEBIT_CREDIT_HEADERS = {"af bij", "bij/af", "debet/credit", "dc", "type"}


@dataclass(frozen=True)
class Transaction:
    row_idx: int
    transaction_date: date
    vendor: str
    amount_cents: int           # signed: negatief = afschrijving (uitgaand)
    currency: str
    description: str | None


def parse_excel(path: Path, *, sheet_name: str | None = None) -> list[Transaction]:
    """Parse afschrijvings-excel. Returns alleen afschrijvingen (uitgaand =
    negatief). Bijschrijvingen worden gefilterd — die hebben geen bon."""
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb[sheet_name] if sheet_name else wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx, headers = _find_header_row(rows)
    if header_idx < 0:
        log.warning("excel %s: kon geen header-rij vinden", path.name)
        return []

    cols = _map_columns(headers)
    if "date" not in cols or "amount" not in cols:
        log.warning("excel %s: ontbreekt date- of amount-kolom (gevonden: %s)",
                     path.name, list(cols.keys()))
        return []

    out: list[Transaction] = []
    for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        try:
            txn = _row_to_transaction(row, cols, row_idx=i)
            if txn is None:
                continue
            # Skip bijschrijvingen — geen bon nodig
            if txn.amount_cents >= 0:
                continue
            out.append(txn)
        except Exception:
            log.exception("excel %s row %d: parse error — skipped",
                           path.name, i)
    return out


def derive_date_window(
    transactions: list[Transaction], *, margin_days: int = 30,
) -> tuple[date, date]:
    """Oudste transactie - margin_days, jongste + margin_days.
    Als de lijst leeg is: returns vandaag±0."""
    if not transactions:
        today = date.today()
        return today, today
    dates = [t.transaction_date for t in transactions]
    from datetime import timedelta
    return min(dates) - timedelta(days=margin_days), max(dates) + timedelta(days=margin_days)


# --- internal helpers ------------------------------------------------------

def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]]:
    """Vind de rij met meeste herkenbare headers. Return (-1, []) als niets
    matcht. Veel exports hebben 1-3 metadata-regels boven de header."""
    best_idx = -1
    best_count = 0
    best_headers: list[str] = []
    for i, row in enumerate(rows[:15]):  # alleen eerste 15 rijen scannen
        cells = [_cell_str(c) for c in row]
        score = _header_score(cells)
        if score > best_count:
            best_count = score
            best_idx = i
            best_headers = cells
    if best_count >= 2:
        return best_idx, best_headers
    return -1, []


def _header_score(cells: list[str]) -> int:
    score = 0
    for c in cells:
        cl = c.strip().lower()
        if cl in _DATE_HEADERS:
            score += 1
        if cl in _AMOUNT_HEADERS or cl in _AMOUNT_OUT_HEADERS or cl in _AMOUNT_IN_HEADERS:
            score += 1
        if cl in _VENDOR_HEADERS or cl in _DESC_HEADERS:
            score += 1
    return score


def _map_columns(headers: list[str]) -> dict[str, int]:
    """Map logical-field-name → column-index. Veld kan ontbreken."""
    cols: dict[str, int] = {}
    for idx, h in enumerate(headers):
        hl = h.strip().lower()
        if "date" not in cols and hl in _DATE_HEADERS:
            cols["date"] = idx
        elif "amount" not in cols and hl in _AMOUNT_HEADERS:
            cols["amount"] = idx
        elif "amount_out" not in cols and hl in _AMOUNT_OUT_HEADERS:
            cols["amount_out"] = idx
        elif "amount_in" not in cols and hl in _AMOUNT_IN_HEADERS:
            cols["amount_in"] = idx
        elif "vendor" not in cols and hl in _VENDOR_HEADERS:
            cols["vendor"] = idx
        elif "desc" not in cols and hl in _DESC_HEADERS:
            cols["desc"] = idx
        elif "debit_credit" not in cols and hl in _DEBIT_CREDIT_HEADERS:
            cols["debit_credit"] = idx
    # Synthesize 'amount' uit gesplitste in/uit kolommen
    if "amount" not in cols and ("amount_out" in cols or "amount_in" in cols):
        cols["amount"] = cols.get("amount_out", cols.get("amount_in", -1))
    return cols


def _row_to_transaction(
    row: tuple[Any, ...], cols: dict[str, int], *, row_idx: int,
) -> Transaction | None:
    if not any(c is not None for c in row):
        return None
    date_val = _parse_date(row[cols["date"]])
    if not date_val:
        return None
    # Gesplitste in/uit kolommen → kies de niet-lege en bepaal sign
    if "amount_out" in cols or "amount_in" in cols:
        out_cents = _parse_amount(row[cols["amount_out"]]) if "amount_out" in cols else None
        in_cents = _parse_amount(row[cols["amount_in"]]) if "amount_in" in cols else None
        if out_cents:
            amount_cents = -abs(out_cents)
        elif in_cents:
            amount_cents = abs(in_cents)
        else:
            return None
    else:
        amount_cents = _parse_amount(row[cols["amount"]])
        if amount_cents is None or amount_cents == 0:
            return None
        if "debit_credit" in cols:
            dc = _cell_str(row[cols["debit_credit"]]).lower()
            if dc in ("af", "debit", "d", "uit"):
                amount_cents = -abs(amount_cents)
            elif dc in ("bij", "credit", "c", "in"):
                amount_cents = abs(amount_cents)
    if amount_cents == 0:
        return None

    vendor = _cell_str(row[cols["vendor"]]).strip() if "vendor" in cols else ""
    desc = _cell_str(row[cols["desc"]]).strip() if "desc" in cols else ""
    if not vendor:
        vendor = _vendor_from_description(desc) if desc else "(unknown)"

    return Transaction(
        row_idx=row_idx,
        transaction_date=date_val,
        vendor=vendor[:120],
        amount_cents=amount_cents,
        currency="EUR",
        description=desc or None,
    )


_IBAN_PREFIX_RE = re.compile(r"^[A-Z]{2}\d{2}[A-Z0-9]{8,30}\s+")
_IBAN_ANY_RE = re.compile(r"\b[A-Z]{2}\d{2}[A-Z0-9]{8,30}\b")
# Words die signaleren dat de vendor-naam VOOR dit woord eindigt
_VENDOR_STOP_WORDS = {"via", "id", "token:", "factuurnummer", "iban",
                       "ref:", "ref.:", "betreft:"}
# Tokens die we negeren als kandidaat (payment-rail noise + algemene woorden)
_CANDIDATE_NOISE = {
    "via", "voor", "voorgeschoten", "voorschot", "id", "ref", "the", "and",
    "of", "te", "een", "het", "de", "van", "factuurnummer", "iban",
    "betreft", "token", "order", "tikkie", "ideal", "sepa", "incasso",
    "stichting", "payments", "payment", "pay", "ccv", "group", "bv", "nv",
    "inc", "llc", "ltd", "gmbh", "bvba", "corp", "co",
    "europe", "international", "intl", "global", "ww", "limited",
    "20-01-2026", "20:56",
}


def _vendor_from_description(desc: str) -> str:
    """Extract de PRIMAIRE vendor uit een vrij-tekst omschrijving.
    Voor multi-candidate extractie: zie `extract_vendor_candidates()`."""
    candidates = extract_vendor_candidates(desc)
    return candidates[0] if candidates else "(unknown)"


def extract_vendor_candidates(desc: str) -> list[str]:
    """Extract MEERDERE vendor-naam kandidaten uit een omschrijving.
    De matcher zoekt parallel op elk en pakt de hoogste-score-hit.

    Voorbeelden:
      'NL13... den Ouden via Tikkie ... Kaartje2go bol voorgeschoten ...'
        → ['den Ouden', 'Kaartje2go', 'bol']
      'LU89... PayPal Europe S.a.r.l. ...'
        → ['PayPal Europe S.a.r.l.', 'PayPal']
      'NL26... Schiphol Parking via CCV Group BV'
        → ['Schiphol Parking']
      'NL51... Porta Sud via Stichting Mollie ...'
        → ['Porta Sud']
    """
    if not desc:
        return []
    s = _IBAN_ANY_RE.sub(" ", desc)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return []

    # Tokenize — accepteer alfanumeriek met internal punctuation
    tokens = s.split()

    # Bouw clusters van consecutive 'goede' tokens (alpha-leading, geen
    # pure-digit, geen ID-look-alike). Stopwords / noise breken cluster.
    clusters: list[list[str]] = []
    current: list[str] = []
    for raw in tokens:
        t = raw.rstrip(":,.;")
        tl = t.lower()
        if not t or len(t) < 3:
            _flush(clusters, current); current = []; continue
        if tl in _CANDIDATE_NOISE:
            _flush(clusters, current); current = []; continue
        if not t[0].isalpha():
            _flush(clusters, current); current = []; continue
        # Look-like-ID: long mixed alphanum (≥6 chars met ≥3 cijfers)
        digit_count = sum(1 for c in t if c.isdigit())
        if len(t) >= 6 and digit_count >= 3:
            _flush(clusters, current); current = []; continue
        current.append(t)
        if len(current) >= 3:
            _flush(clusters, current); current = []
    _flush(clusters, current)

    # Bouw kandidaten: full cluster + (als multi-word) single first-word
    candidates: list[str] = []
    seen: set[str] = set()
    for cluster in clusters:
        full = " ".join(cluster)
        key = full.lower()
        if key not in seen:
            candidates.append(full)
            seen.add(key)
        if len(cluster) > 1:
            single = cluster[0]
            sk = single.lower()
            if sk not in seen and single.lower() not in _CANDIDATE_NOISE:
                candidates.append(single)
                seen.add(sk)
    return candidates[:6]


def _flush(clusters: list[list[str]], current: list[str]) -> None:
    if current:
        clusters.append(current.copy())


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d",
                 "%d.%m.%Y", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


_AMOUNT_RE = re.compile(r"[^\d,.\-]")


def _parse_amount(value: Any) -> int | None:
    """Naar cents. Accepteert NL/EN notatie: 1.234,56 of 1,234.56 of 1234.56.
    Returns None als niet leesbaar."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(round(float(value) * 100))
    s = _AMOUNT_RE.sub("", str(value).strip())
    if not s or s in ("-", "."):
        return None
    # NL-notatie: punt = duizendtal, komma = decimaal
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return int(round(float(s) * 100))
    except ValueError:
        return None


def _cell_str(cell: Any) -> str:
    if cell is None:
        return ""
    return str(cell)
