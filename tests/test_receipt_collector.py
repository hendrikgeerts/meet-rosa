"""Tests voor receipt-collector: schema, parser, matcher-scoring, runner."""
from __future__ import annotations

import sqlite3
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import pytest

from extensions.receipt_collector.matcher import (
    Attachment, MatchCandidate, _amount_search_strings, score_candidate,
)
from extensions.receipt_collector.parser import (
    Transaction, derive_date_window, parse_excel,
    _parse_amount, _parse_date, _vendor_from_description,
)
from extensions.receipt_collector.runner import (
    _build_search_vendors, _is_reverse_match_eligible, run_receipt_collection,
)
from extensions.receipt_collector.schema import (
    find_vendor_strategy, init_receipt_collector_schema,
    insert_run, insert_run_item, list_run_items, list_runs,
    list_vendor_strategies, upsert_vendor_strategy,
    update_run_counts, update_run_item,
)
from extensions.receipt_collector.tools import (
    receipt_run_start_handler, receipt_run_status_handler,
    receipt_runs_list_handler, vendor_strategies_list_handler,
    vendor_strategy_remember_handler,
)


@pytest.fixture
def db(tmp_path: Path) -> Path:
    p = tmp_path / "rc.db"
    init_receipt_collector_schema(p)
    return p


# --- schema --------------------------------------------------------------

# --- HIGH-4: receipt_run_start path-traversal guard ---------------------

def test_receipt_run_start_rejects_path_outside_receipts_root(db: Path, tmp_path: Path) -> None:
    """ISO_AUDIT 2026-05 HIGH-4: Claude-supplied excel_path must not be
    able to point outside ~/PA-Receipts/."""
    receipts_root = tmp_path / "receipts"
    receipts_root.mkdir()
    out = receipt_run_start_handler(
        db, {"excel_path": "/etc/passwd"}, output_root=receipts_root,
    )
    assert "error" in out
    assert "within" in out["error"].lower() or "PA-Receipts" in out["error"]


def test_receipt_run_start_rejects_traversal(db: Path, tmp_path: Path) -> None:
    """Relative paths with `..` resolved → not under receipts_root → reject."""
    receipts_root = tmp_path / "receipts"
    (receipts_root / "inbox").mkdir(parents=True)
    out = receipt_run_start_handler(
        db, {"excel_path": "../../etc/passwd"}, output_root=receipts_root,
    )
    assert "error" in out


def test_receipt_run_start_rejects_non_excel_extension(db: Path, tmp_path: Path) -> None:
    receipts_root = tmp_path / "receipts"
    inbox = receipts_root / "inbox"
    inbox.mkdir(parents=True)
    bogus = inbox / "evil.sh"
    bogus.write_text("#!/bin/sh\nrm -rf /\n")
    out = receipt_run_start_handler(
        db, {"excel_path": "evil.sh"}, output_root=receipts_root,
    )
    assert "error" in out
    assert ".xlsx" in out["error"] or ".xls" in out["error"]


def test_receipt_run_start_rejects_nonexistent(db: Path, tmp_path: Path) -> None:
    receipts_root = tmp_path / "receipts"
    (receipts_root / "inbox").mkdir(parents=True)
    out = receipt_run_start_handler(
        db, {"excel_path": "does-not-exist.xlsx"}, output_root=receipts_root,
    )
    assert "error" in out
    assert "not found" in out["error"]


def test_insert_run_and_items(db: Path) -> None:
    with sqlite3.connect(db) as c:
        rid = insert_run(
            c, excel_path="/tmp/q2.xlsx", output_dir="/tmp/run",
            period_label="Q2-2026", date_window_start=0,
            date_window_end=100, transaction_count=2,
        )
        assert rid > 0
        iid1 = insert_run_item(
            c, run_id=rid, row_idx=2, transaction_date=10,
            vendor_raw="AWS", amount_cents=-12750,
        )
        iid2 = insert_run_item(
            c, run_id=rid, row_idx=3, transaction_date=20,
            vendor_raw="Bol.com", amount_cents=-3499,
        )
    assert iid1 > 0 and iid2 > 0

    with sqlite3.connect(db) as c:
        items = list_run_items(c, rid)
    assert len(items) == 2
    assert items[0]["vendor_raw"] == "AWS"


def test_update_run_counts_and_status(db: Path) -> None:
    with sqlite3.connect(db) as c:
        rid = insert_run(c, excel_path="x", output_dir="y",
                          period_label=None, date_window_start=0,
                          date_window_end=0, transaction_count=5)
        update_run_counts(
            c, rid, matched=3, needs_portal=1, unknown=1,
            status="needs_input", completed=False,
        )
        row = c.execute("SELECT status, matched_count FROM receipt_runs WHERE id=?",
                          (rid,)).fetchone()
    assert row[0] == "needs_input"
    assert row[1] == 3


def test_upsert_and_find_vendor_by_alias(db: Path) -> None:
    with sqlite3.connect(db) as c:
        upsert_vendor_strategy(
            c, name="Amazon Web Services",
            source_kind="email",
            aliases=["aws", "amazon ws"],
            email_query_hint="from:billing@aws.amazon.com",
        )
        v = find_vendor_strategy(c, vendor_text="AWS Inc")
    assert v is not None
    assert v["name"] == "Amazon Web Services"
    assert v["source_kind"] == "email"


def test_upsert_idempotent(db: Path) -> None:
    with sqlite3.connect(db) as c:
        upsert_vendor_strategy(c, name="X", source_kind="email")
        upsert_vendor_strategy(c, name="X", source_kind="portal",
                                portal_url="https://x.test")
        rows = list_vendor_strategies(c)
    assert len(rows) == 1
    assert rows[0]["source_kind"] == "portal"


def test_invalid_source_kind(db: Path) -> None:
    with sqlite3.connect(db) as c:
        with pytest.raises(ValueError):
            upsert_vendor_strategy(c, name="X", source_kind="weird")


# --- parser --------------------------------------------------------------

def test_parse_amount_nl_format() -> None:
    assert _parse_amount("1.234,56") == 123456
    assert _parse_amount("12,50") == 1250
    assert _parse_amount("-99,99") == -9999


def test_parse_amount_en_format() -> None:
    assert _parse_amount("1,234.56") == 123456
    assert _parse_amount("12.50") == 1250


def test_parse_amount_handles_currency_symbols() -> None:
    assert _parse_amount("€ 12,50") == 1250
    assert _parse_amount("EUR 12.50") == 1250


def test_parse_amount_invalid_returns_none() -> None:
    assert _parse_amount("") is None
    assert _parse_amount("nope") is None
    assert _parse_amount(None) is None


def test_parse_date_formats() -> None:
    assert _parse_date("2026-04-27") == date(2026, 4, 27)
    assert _parse_date("27-04-2026") == date(2026, 4, 27)
    assert _parse_date("27/04/2026") == date(2026, 4, 27)


def test_vendor_from_description_strips_iban() -> None:
    # PayPal: "Europe" is in noise-list dus eerste candidate is "PayPal"
    assert _vendor_from_description(
        "LU89751000135104200E PayPal Europe S.a.r.l. et Cie ..."
    ) == "PayPal"
    assert _vendor_from_description(
        "NL13ABNA0506417344 den Ouden via Tikkie 001160770007"
    ) == "den Ouden"
    assert _vendor_from_description(
        "NL26INGB0006865503 Schiphol Parking via CCV"
    ) == "Schiphol Parking"
    assert _vendor_from_description(
        "NL51DEUT0265262461 Porta Sud via Stichting Mollie"
    ) == "Porta Sud"


def test_vendor_from_description_no_iban() -> None:
    # BABYLOVEGROWTH heeft geen IBAN-prefix → eerste cluster van 3 woorden
    assert _vendor_from_description(
        "BABYLOVEGROWTH SAN RAFAEL USACA Token: 5xxxx2394"
    ) == "BABYLOVEGROWTH SAN RAFAEL"


def test_vendor_from_description_empty() -> None:
    assert _vendor_from_description("") == "(unknown)"
    assert _vendor_from_description("   ") == "(unknown)"


def test_extract_multiple_candidates_kaartje2go() -> None:
    """Use-case: voorgeschoten via Tikkie → echte vendor staat verderop."""
    from extensions.receipt_collector.parser import extract_vendor_candidates
    cands = extract_vendor_candidates(
        "NL13ABNA0506417344 den Ouden via Tikkie 001160770007 0031645686029196 "
        "Kaartje2go bol voorgeschoten NL14ABNA0438180445"
    )
    # Both 'den Ouden' (intermediair) AND 'Kaartje2go' (echte vendor) komen mee
    assert "den Ouden" in cands
    assert "Kaartje2go" in cands or "Kaartje2go bol" in cands


def test_extract_candidates_drops_iban_and_ids() -> None:
    from extensions.receipt_collector.parser import extract_vendor_candidates
    cands = extract_vendor_candidates(
        "NL51DEUT0265262461 Porta Sud via Stichting Mollie Payments "
        "dade9d38e8c6ee9efb0f4ad98d36a0a6 8152605664655444"
    )
    assert "Porta Sud" in cands
    # IBAN/IDs niet in kandidaten
    assert not any("DEUT" in c or "dade9d38" in c for c in cands)


def test_extract_candidates_empty() -> None:
    from extensions.receipt_collector.parser import extract_vendor_candidates
    assert extract_vendor_candidates("") == []
    assert extract_vendor_candidates("   ") == []


def test_parse_excel_real_file(tmp_path: Path) -> None:
    """Maak een realistisch excel + parse 'm."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Bank afschriften Q2 2026"])  # metadata-rij
    ws.append([])
    ws.append(["Datum", "Naam", "Bedrag", "Omschrijving"])
    ws.append([date(2026, 4, 5), "AWS",      "-127,50", "Maandelijkse hosting"])
    ws.append([date(2026, 4, 12), "Bol.com", "-34,99",  "Boek"])
    ws.append([date(2026, 4, 15), "Loonbetaling", "5000,00", "Salaris"])  # bijschrijving — moet weggefilterd worden
    ws.append([date(2026, 4, 20), "Microsoft", "-21,00", "M365 license"])
    excel_path = tmp_path / "q2.xlsx"
    wb.save(excel_path)

    txns = parse_excel(excel_path)
    assert len(txns) == 3
    assert all(t.amount_cents < 0 for t in txns)
    assert {t.vendor for t in txns} == {"AWS", "Bol.com", "Microsoft"}
    aws = next(t for t in txns if t.vendor == "AWS")
    assert aws.amount_cents == -12750
    assert aws.transaction_date == date(2026, 4, 5)


def test_parse_excel_no_header(tmp_path: Path) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["random", "garbage", "stuff"])
    ws.append(["a", "b", "c"])
    excel_path = tmp_path / "bad.xlsx"
    wb.save(excel_path)
    assert parse_excel(excel_path) == []


def test_derive_date_window() -> None:
    txns = [
        Transaction(row_idx=1, transaction_date=date(2026, 4, 5),
                     vendor="A", amount_cents=-100, currency="EUR",
                     description=None),
        Transaction(row_idx=2, transaction_date=date(2026, 4, 28),
                     vendor="B", amount_cents=-200, currency="EUR",
                     description=None),
    ]
    start, end = derive_date_window(txns, margin_days=30)
    assert start == date(2026, 3, 6)
    assert end == date(2026, 5, 28)


def test_derive_date_window_empty() -> None:
    start, end = derive_date_window([])
    assert start == end == date.today()


# --- matcher scoring ------------------------------------------------------

def _txn(amount: int = -12750, vendor: str = "AWS",
         d: date = date(2026, 4, 5)) -> Transaction:
    return Transaction(row_idx=1, transaction_date=d, vendor=vendor,
                        amount_cents=amount, currency="EUR", description=None)


def _make_pdf_with_amount(amount_text: str) -> bytes:
    """Mini-PDF die amount_text in extract_text terugbrengt."""
    from pypdf import PdfWriter
    from pypdf.generic import (
        ArrayObject, DictionaryObject, NameObject, NumberObject,
        TextStringObject,
    )
    # Simpel: schrijf een lege pdf en voeg een visible-text annotation toe.
    # Voor de test maken we een page met een Contents stream die 'amount_text'
    # in een BT...ET blok zet.
    writer = PdfWriter()
    page = writer.add_blank_page(width=200, height=200)
    # pypdf heeft geen native draw_text; gebruik reportlab-style hack via
    # _content_streams. Voor onze test: monkeypatch werkt makkelijker.
    # → gebruik simpele PDF-byte string ipv real PDF
    raw = (
        b"%PDF-1.4\n"
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n"
        b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n"
        b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 200 200] "
        b"/Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >> endobj\n"
        b"4 0 obj << /Length 44 >> stream\n"
        b"BT /F1 12 Tf 10 100 Td (" + amount_text.encode() + b") Tj ET\n"
        b"endstream endobj\n"
        b"5 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n"
        b"xref\n0 6\n0000000000 65535 f\n"
        b"0000000010 00000 n\n0000000053 00000 n\n0000000098 00000 n\n"
        b"0000000180 00000 n\n0000000260 00000 n\n"
        b"trailer << /Size 6 /Root 1 0 R >>\nstartxref\n340\n%%EOF\n"
    )
    return raw


def test_score_candidate_amount_match() -> None:
    pdf = _make_pdf_with_amount("127,50")
    cand = MatchCandidate(
        source="gmail", message_id="x",
        from_addr="billing@aws.amazon.com", subject="Invoice AWS",
        occurred_at=date(2026, 4, 5),
        attachments=[Attachment(filename="invoice.pdf",
                                  mime_type="application/pdf", data=pdf)],
    )
    score, reasons = score_candidate(cand, _txn())
    # date+0.30 + vendor+0.20 + amount+0.50 = 1.0
    assert score >= 0.9
    assert any("amount" in r for r in reasons)


def test_score_candidate_no_amount_partial_score() -> None:
    cand = MatchCandidate(
        source="gmail", message_id="x",
        from_addr="some@other.com", subject="Order shipped",
        occurred_at=date(2026, 4, 12),
        attachments=[Attachment(filename="x.pdf",
                                  mime_type="application/pdf",
                                  data=b"%PDF-1.4 not-a-real-pdf")],
    )
    score, _ = score_candidate(cand, _txn())
    # alleen date-score (~7d off → 0.15), geen vendor, geen amount
    assert score < 0.5


def test_score_candidate_far_date_low_score() -> None:
    cand = MatchCandidate(
        source="gmail", message_id="x",
        from_addr="billing@aws.amazon.com", subject="x",
        occurred_at=date(2026, 1, 1),  # >30d off
        attachments=[Attachment(filename="x.pdf",
                                  mime_type="application/pdf",
                                  data=b"empty")],
    )
    score, _ = score_candidate(cand, _txn())
    assert score < 0.5  # alleen vendor-score van 0.2


# --- runner --------------------------------------------------------------

def _make_excel(tmp_path: Path) -> Path:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Datum", "Naam", "Bedrag"])
    ws.append([date(2026, 4, 5), "AWS", "-127,50"])
    p = tmp_path / "test.xlsx"
    wb.save(p)
    return p


def test_runner_no_sources_marks_unknown(db: Path, tmp_path: Path) -> None:
    excel = _make_excel(tmp_path)
    out = tmp_path / "out"
    summary = run_receipt_collection(
        excel_path=excel, db_path=db, output_root=out,
        gmail=None, imap_accounts=[],
    )
    assert summary["transaction_count"] == 1
    assert summary["matched"] == 0
    assert summary["unknown_vendor"] == 1
    assert summary["status"] == "needs_input"


def test_runner_portal_strategy_skips_search(db: Path, tmp_path: Path) -> None:
    with sqlite3.connect(db) as c:
        upsert_vendor_strategy(
            c, name="AWS", source_kind="portal",
            aliases=["aws"],
            portal_url="https://console.aws.amazon.com",
            portal_notes="Login → Billing → Invoices",
        )
    excel = _make_excel(tmp_path)
    out = tmp_path / "out"
    summary = run_receipt_collection(
        excel_path=excel, db_path=db, output_root=out,
        gmail=None, imap_accounts=[],
    )
    assert summary["needs_portal"] == 1
    assert summary["matched"] == 0


# --- always-on Ollama vendor extraction ------------------------------------

def test_build_search_vendors_always_calls_ollama_with_description() -> None:
    """Regex-extract levert ≥2 kandidaten op, maar Ollama-pass moet
    nog steeds draaien — die vangt subtiele cases (voorgeschoten X)."""
    txn = Transaction(
        row_idx=1, transaction_date=date(2026, 4, 5),
        vendor="50456 - Grammarly Inc.",
        amount_cents=-14400, currency="EUR",
        description="GRAMMARLY CO*RIQXLRL SAN FRANCISCOUSACA",
    )
    fake_ollama = MagicMock()
    fake_response = MagicMock()
    fake_block = MagicMock()
    fake_block.type = "text"
    fake_block.text = '["Grammarly", "Chargebee"]'
    fake_response.content = [fake_block]
    fake_ollama.chat.return_value = fake_response

    vendors = _build_search_vendors(txn, ollama=fake_ollama)

    # Ollama-pass is altijd uitgevoerd (gate stond eerst op ≤1 candidate)
    fake_ollama.chat.assert_called_once()
    # 'Chargebee' kwam niet uit regex — alleen via Ollama. Bewijs dat de
    # always-on pass z'n werk doet.
    assert "Chargebee" in vendors


def test_build_search_vendors_no_description_skips_ollama() -> None:
    """Geen description → niets om uit te halen, geen ollama-call."""
    txn = Transaction(
        row_idx=1, transaction_date=date(2026, 4, 5),
        vendor="AWS", amount_cents=-12750, currency="EUR",
        description=None,
    )
    fake_ollama = MagicMock()
    _build_search_vendors(txn, ollama=fake_ollama)
    fake_ollama.chat.assert_not_called()


# --- reverse-match (amount-first) -----------------------------------------

def test_clean_vendor_strips_mollie_prefix() -> None:
    """Mol* is Mollie's merchant-tag — strip net als 'paypal *'."""
    from extensions.receipt_collector.matcher import _clean_vendor_for_search
    assert _clean_vendor_for_search("Mol*ADL Video B V") == "ADL Video"
    assert _clean_vendor_for_search("Mol*Stichting Paydr Amstelveen") == "Stichting Paydr"


def test_amount_search_strings_nl_en() -> None:
    assert _amount_search_strings(12750) == ["127,50", "127.50"]
    assert _amount_search_strings(-12750) == ["127,50", "127.50"]
    assert _amount_search_strings(0) == []


def test_reverse_match_eligibility() -> None:
    base = Transaction(row_idx=1, transaction_date=date(2026, 4, 5),
                        vendor="X", amount_cents=-12750, currency="EUR",
                        description=None)
    assert _is_reverse_match_eligible(base) is True
    # Te klein
    assert _is_reverse_match_eligible(
        Transaction(row_idx=1, transaction_date=date(2026, 4, 5),
                     vendor="X", amount_cents=-450, currency="EUR",
                     description=None)) is False
    # Rond bedrag (geen cents)
    assert _is_reverse_match_eligible(
        Transaction(row_idx=1, transaction_date=date(2026, 4, 5),
                     vendor="X", amount_cents=-5000, currency="EUR",
                     description=None)) is False


def test_runner_reverse_match_recovers_unmatched(
    db: Path, tmp_path: Path,
) -> None:
    """Vendor-search vindt niets, maar amount-search vindt een mail met
    PDF waar het bedrag exact in staat. Reverse-pass moet dit oppikken."""
    excel = _make_excel(tmp_path)  # AWS, -127,50, 2026-04-05
    pdf = _make_pdf_with_amount("127,50")

    fake_gmail = MagicMock()

    def search_side_effect(query: str, max_results: int = 10) -> list[dict]:
        # Vendor-search "AWS" → niks
        if "AWS" in query or "aws" in query:
            return []
        # Amount-search "127,50" of "127.50" → één hit
        if "127,50" in query or "127.50" in query:
            return [{"id": "msg-rev", "thread_id": "t-rev"}]
        return []

    fake_gmail.search.side_effect = search_side_effect
    fake_gmail.get_message_full.return_value = {
        "id": "msg-rev",
        "payload": {
            "headers": [
                # Note: vendor 'AWS' staat NIET in from/subject
                {"name": "From", "value": "billing@chargebee.com"},
                {"name": "Subject", "value": "Receipt #12345"},
                {"name": "Date", "value": "Mon, 05 Apr 2026 09:00:00 +0000"},
            ],
            "parts": [
                {"filename": "receipt.pdf", "mimeType": "application/pdf",
                 "body": {"attachmentId": "att-rev"}},
            ],
        },
    }
    fake_gmail.get_attachment.return_value = pdf

    out = tmp_path / "out"
    summary = run_receipt_collection(
        excel_path=excel, db_path=db, output_root=out,
        gmail=fake_gmail, imap_accounts=[],
    )
    assert summary["matched"] == 1
    assert summary["unknown_vendor"] == 0
    # Note must flag this as review
    with sqlite3.connect(db) as c:
        row = c.execute(
            "SELECT notes FROM receipt_run_items WHERE run_id=?",
            (summary["run_id"],)).fetchone()
    assert row is not None
    assert "amount-only match (review)" in (row[0] or "")


def test_runner_reverse_match_skipped_for_round_amount(
    db: Path, tmp_path: Path,
) -> None:
    """€50,00 is te gangbaar — reverse-pass moet eligibility-check skippen."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Datum", "Naam", "Bedrag"])
    ws.append([date(2026, 4, 5), "Onbekend", "-50,00"])
    excel = tmp_path / "round.xlsx"
    wb.save(excel)
    pdf = _make_pdf_with_amount("50,00")

    fake_gmail = MagicMock()
    fake_gmail.search.return_value = [{"id": "msg-1", "thread_id": "t-1"}]
    fake_gmail.get_message_full.return_value = {
        "id": "msg-1",
        "payload": {
            "headers": [
                {"name": "From", "value": "x@y.com"},
                {"name": "Subject", "value": "Iets"},
                {"name": "Date", "value": "Mon, 05 Apr 2026 09:00:00 +0000"},
            ],
            "parts": [
                {"filename": "f.pdf", "mimeType": "application/pdf",
                 "body": {"attachmentId": "a-1"}},
            ],
        },
    }
    fake_gmail.get_attachment.return_value = pdf

    out = tmp_path / "out"
    summary = run_receipt_collection(
        excel_path=excel, db_path=db, output_root=out,
        gmail=fake_gmail, imap_accounts=[],
    )
    # Eerste pass scoort wel (vendor "Onbekend" niet in subject, maar
    # amount + datum = 0.5 + 0.3 = 0.8 ≥ 0.55) — dus wordt matched.
    # Maar de reverse-pass MAG niet eens triggeren want round amount.
    # Hier checken we vooral dat er geen extra reverse-call is gedaan.
    # search() wordt 1x aangeroepen (vendor-pass), niet nogmaals voor amount.
    queries = [c.kwargs.get("query") or c.args[0]
                for c in fake_gmail.search.call_args_list
                if c.args or c.kwargs]
    # Geen query mag puur een amount-string als enige search-term hebben
    amount_only_queries = [q for q in queries
                            if q and '"50,00"' in q and "Onbekend" not in q]
    assert amount_only_queries == []


# --- email_to_pdf -------------------------------------------------------

def test_email_to_pdf_looks_like_invoice() -> None:
    from extensions.receipt_collector.email_to_pdf import looks_like_invoice
    assert looks_like_invoice("Your invoice for April") is True
    assert looks_like_invoice("Receipt for €127.50") is True
    assert looks_like_invoice("Order shipped, tracking #X") is False
    assert looks_like_invoice("") is False


def test_email_to_pdf_render_roundtrip() -> None:
    """Gerenderde PDF moet bedrag bevatten zodat _amount_in_pdf 'm vindt."""
    from io import BytesIO

    from extensions.receipt_collector.email_to_pdf import render_email_as_pdf
    from pypdf import PdfReader

    pdf = render_email_as_pdf(
        headers={"From": "billing@datadog.com", "Subject": "Invoice"},
        body_text="Total amount: 275.46 EUR\nThank you for your business.",
    )
    assert pdf is not None
    assert pdf[:5] == b"%PDF-"
    text = PdfReader(BytesIO(pdf)).pages[0].extract_text()
    assert "275.46" in text
    assert "billing@datadog.com" in text


def test_email_to_pdf_empty_returns_none() -> None:
    from extensions.receipt_collector.email_to_pdf import render_email_as_pdf
    assert render_email_as_pdf(headers={}, body_text=None, body_html=None) is None
    assert render_email_as_pdf(headers={}, body_text="   ", body_html="") is None


def test_runner_renders_email_pdf_when_no_attachment(
    db: Path, tmp_path: Path,
) -> None:
    """Mail zonder PDF-attachment maar met invoice-keywords + bedrag in
    body → Rosa rendert evidence-PDF en matcht alsnog."""
    excel = _make_excel(tmp_path)  # AWS, -127,50, 2026-04-05

    fake_gmail = MagicMock()
    fake_gmail.search.return_value = [{"id": "msg-html", "thread_id": "t"}]
    # Mail met text/plain body waarin invoice-keywords + amount.
    # Geen filename/attachmentId → geen 'echte' PDF, dwingt body-render.
    import base64
    body_text = "Invoice for your subscription\nAmount: 127,50 EUR\nThank you."
    encoded = base64.urlsafe_b64encode(body_text.encode()).decode().rstrip("=")
    fake_gmail.get_message_full.return_value = {
        "id": "msg-html",
        "payload": {
            "headers": [
                {"name": "From", "value": "billing@netsuite.com"},
                {"name": "Subject", "value": "Your invoice"},
                {"name": "Date", "value": "Mon, 05 Apr 2026 09:00:00 +0000"},
            ],
            "mimeType": "text/plain",
            "body": {"data": encoded},
        },
    }

    out = tmp_path / "out"
    summary = run_receipt_collection(
        excel_path=excel, db_path=db, output_root=out,
        gmail=fake_gmail, imap_accounts=[],
    )
    assert summary["matched"] == 1
    # Saved attachment is an email-evidence PDF
    run_dir = Path(summary["output_dir"])
    files = list(run_dir.iterdir())
    assert len(files) == 1
    assert files[0].suffix == ".pdf"


def test_runner_skips_render_for_non_invoice_email(
    db: Path, tmp_path: Path,
) -> None:
    """Mail zonder PDF en zonder invoice-keywords → geen evidence-PDF,
    geen match (oude gedrag behouden voor non-invoice mails)."""
    excel = _make_excel(tmp_path)
    import base64
    body_text = "Order shipped, tracking #12345"
    encoded = base64.urlsafe_b64encode(body_text.encode()).decode().rstrip("=")

    fake_gmail = MagicMock()
    fake_gmail.search.return_value = [{"id": "msg-x"}]
    fake_gmail.get_message_full.return_value = {
        "id": "msg-x",
        "payload": {
            "headers": [
                {"name": "From", "value": "shop@example.com"},
                {"name": "Subject", "value": "Your order shipped"},
                {"name": "Date", "value": "Mon, 05 Apr 2026 09:00:00 +0000"},
            ],
            "mimeType": "text/plain",
            "body": {"data": encoded},
        },
    }

    out = tmp_path / "out"
    summary = run_receipt_collection(
        excel_path=excel, db_path=db, output_root=out,
        gmail=fake_gmail, imap_accounts=[],
    )
    assert summary["matched"] == 0
    assert summary["unknown_vendor"] == 1


def test_runner_ignore_strategy_marks_ignored(
    db: Path, tmp_path: Path,
) -> None:
    """source_kind='ignore' → status='ignored', geen mail-search."""
    with sqlite3.connect(db) as c:
        upsert_vendor_strategy(
            c, name="ADL Video test-sub", source_kind="ignore",
            aliases=["AWS"],
            portal_notes="test subscription, niet relevant",
        )
    excel = _make_excel(tmp_path)
    fake_gmail = MagicMock()

    out = tmp_path / "out"
    summary = run_receipt_collection(
        excel_path=excel, db_path=db, output_root=out,
        gmail=fake_gmail, imap_accounts=[],
    )
    assert summary["ignored"] == 1
    assert summary["matched"] == 0
    assert summary["unknown_vendor"] == 0
    fake_gmail.search.assert_not_called()


def test_runner_physical_strategy_marks_physical_only(
    db: Path, tmp_path: Path,
) -> None:
    """source_kind='physical' → status='physical_only', geen mail-search."""
    with sqlite3.connect(db) as c:
        upsert_vendor_strategy(
            c, name="AWS pin", source_kind="physical",
            aliases=["AWS"],
        )
    excel = _make_excel(tmp_path)
    fake_gmail = MagicMock()

    out = tmp_path / "out"
    summary = run_receipt_collection(
        excel_path=excel, db_path=db, output_root=out,
        gmail=fake_gmail, imap_accounts=[],
    )
    assert summary["physical_only"] == 1
    assert summary["matched"] == 0
    fake_gmail.search.assert_not_called()


def test_runner_dedup_message_id_across_txns(
    db: Path, tmp_path: Path,
) -> None:
    """Twee transacties op zelfde datum met identiek bedrag mogen niet
    dezelfde mail krijgen toegewezen (Zoomash-bug uit Run-7)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Datum", "Naam", "Bedrag"])
    ws.append([date(2026, 4, 5), "AWS", "-127,50"])  # txn 1
    ws.append([date(2026, 4, 5), "AWS", "-127,50"])  # txn 2 — duplicate
    excel = tmp_path / "dup.xlsx"
    wb.save(excel)
    pdf = _make_pdf_with_amount("127,50")

    fake_gmail = MagicMock()
    fake_gmail.search.return_value = [{"id": "msg-1", "thread_id": "t-1"}]
    fake_gmail.get_message_full.return_value = {
        "id": "msg-1",
        "payload": {
            "headers": [
                {"name": "From", "value": "billing@aws.amazon.com"},
                {"name": "Subject", "value": "AWS Invoice"},
                {"name": "Date", "value": "Mon, 05 Apr 2026 09:00:00 +0000"},
            ],
            "parts": [
                {"filename": "i.pdf", "mimeType": "application/pdf",
                 "body": {"attachmentId": "att"}},
            ],
        },
    }
    fake_gmail.get_attachment.return_value = pdf

    out = tmp_path / "out"
    summary = run_receipt_collection(
        excel_path=excel, db_path=db, output_root=out,
        gmail=fake_gmail, imap_accounts=[],
    )
    # Eerste txn matched op msg-1, tweede mag dezelfde msg-1 niet pakken
    assert summary["matched"] == 1
    with sqlite3.connect(db) as c:
        rows = c.execute(
            "SELECT status, source_message_id FROM receipt_run_items "
            "WHERE run_id=? ORDER BY id", (summary["run_id"],)).fetchall()
    statuses = [r[0] for r in rows]
    msg_ids = [r[1] for r in rows if r[1]]
    # Geen 2 items met dezelfde source_message_id
    assert len(msg_ids) == len(set(msg_ids))
    assert statuses.count("matched") == 1


def test_runner_portal_with_email_hint_tries_mail_first(
    db: Path, tmp_path: Path,
) -> None:
    """Portal + email_query_hint: probeer mail eerst, niet alleen portal-skip."""
    with sqlite3.connect(db) as c:
        upsert_vendor_strategy(
            c, name="AWS", source_kind="portal",
            aliases=["aws"],
            email_query_hint="from:billing@aws.amazon.com",
            portal_url="https://console.aws.amazon.com",
        )
    excel = _make_excel(tmp_path)
    pdf = _make_pdf_with_amount("127,50")

    fake_gmail = MagicMock()
    fake_gmail.search.return_value = [{"id": "msg-aws", "thread_id": "t-aws"}]
    fake_gmail.get_message_full.return_value = {
        "id": "msg-aws",
        "payload": {
            "headers": [
                {"name": "From", "value": "billing@aws.amazon.com"},
                {"name": "Subject", "value": "AWS Invoice"},
                {"name": "Date", "value": "Mon, 05 Apr 2026 09:00:00 +0000"},
            ],
            "parts": [
                {"filename": "invoice.pdf", "mimeType": "application/pdf",
                 "body": {"attachmentId": "att"}},
            ],
        },
    }
    fake_gmail.get_attachment.return_value = pdf

    out = tmp_path / "out"
    summary = run_receipt_collection(
        excel_path=excel, db_path=db, output_root=out,
        gmail=fake_gmail, imap_accounts=[],
    )
    # Mail-pad gebruikt → matched, niet needs_portal
    assert summary["matched"] == 1
    assert summary["needs_portal"] == 0


def test_runner_with_gmail_match(db: Path, tmp_path: Path) -> None:
    excel = _make_excel(tmp_path)
    pdf = _make_pdf_with_amount("127,50")

    fake_gmail = MagicMock()
    fake_gmail.search.return_value = [
        {"id": "msg-1", "thread_id": "t-1"},
    ]
    fake_gmail.get_message_full.return_value = {
        "id": "msg-1",
        "payload": {
            "headers": [
                {"name": "From", "value": "billing@aws.amazon.com"},
                {"name": "Subject", "value": "AWS Invoice April"},
                {"name": "Date", "value": "Mon, 05 Apr 2026 09:00:00 +0000"},
            ],
            "parts": [
                {"filename": "invoice.pdf", "mimeType": "application/pdf",
                 "body": {"attachmentId": "att-1"}},
            ],
        },
    }
    fake_gmail.get_attachment.return_value = pdf

    out = tmp_path / "out"
    summary = run_receipt_collection(
        excel_path=excel, db_path=db, output_root=out,
        gmail=fake_gmail, imap_accounts=[],
    )
    assert summary["matched"] == 1
    assert summary["status"] == "completed"
    # Attachment is opgeslagen
    run_dir = Path(summary["output_dir"])
    files = list(run_dir.iterdir())
    assert len(files) == 1
    assert files[0].suffix == ".pdf"


# --- tools ---------------------------------------------------------------

def test_vendor_strategy_remember_tool(db: Path) -> None:
    out = vendor_strategy_remember_handler(db, {
        "name": "Microsoft", "source_kind": "portal",
        "portal_url": "https://admin.microsoft.com",
        "portal_notes": "Billing → Invoices",
    })
    assert out["ok"] is True
    listed = vendor_strategies_list_handler(db, {})
    assert len(listed) == 1
    assert listed[0]["name"] == "Microsoft"


def test_vendor_strategy_invalid_source_kind(db: Path) -> None:
    out = vendor_strategy_remember_handler(db, {
        "name": "X", "source_kind": "ufo",
    })
    assert "error" in out


def test_runs_list_tool(db: Path) -> None:
    with sqlite3.connect(db) as c:
        insert_run(c, excel_path="x", output_dir="y", period_label="Q1-2026",
                    date_window_start=0, date_window_end=0,
                    transaction_count=3)
    rows = receipt_runs_list_handler(db, {"limit": 5})
    assert len(rows) == 1
    assert rows[0]["period_label"] == "Q1-2026"


def test_run_status_unknown(db: Path) -> None:
    out = receipt_run_status_handler(db, {"run_id": 999})
    assert "error" in out


def test_run_status_returns_items(db: Path) -> None:
    with sqlite3.connect(db) as c:
        rid = insert_run(c, excel_path="x", output_dir="y",
                          period_label=None, date_window_start=0,
                          date_window_end=0, transaction_count=1)
        insert_run_item(c, run_id=rid, row_idx=2, transaction_date=10,
                         vendor_raw="AWS", amount_cents=-12750)
    out = receipt_run_status_handler(db, {"run_id": rid})
    assert out["run"]["transaction_count"] == 1
    assert len(out["items"]) == 1
    assert out["items"][0]["amount_eur"] == -127.50
